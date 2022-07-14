# -*- coding: utf-8 -*-

raise IOError("Don't execute me anymore!!!!")

import io
import os
from os import listdir
from os.path import isfile, join
import openpyxl
from openpyxl.utils import get_column_letter

from django.db import IntegrityError
from django.core.files import File
from django.core.management.base import BaseCommand
from django.conf import settings
from django.utils.translation import ugettext_lazy as _

from codenerix.helpers import nameunify
from codenerix.lib.debugger import Debugger
from codenerix_products.models import TypeTax, Family, Category, Subcategory, MODELS_SLUG, MODELS_BRANDS, MODELS_PRODUCTS, MODELS_PRODUCTS_FINAL, Product, ProductFinal, Attribute, MODELS, TYPE_VALUE_LIST, GroupValueAttribute, OptionValueAttribute, ProductFinalImage, ProductFinalAttribute, ProductImage, ProductUnique
for info in MODELS_SLUG + MODELS_BRANDS + MODELS_PRODUCTS + MODELS + MODELS_PRODUCTS_FINAL:
    field = info[0]
    model = info[1]
    for lang_code in settings.LANGUAGES_DATABASES:
        query = "from codenerix_products.models import {}Text{}\n".format(model, lang_code)
        print(query)
        exec(query)


class Command(BaseCommand, Debugger):
    PROJECT = 'urbangest'
    PATH = 'common/management/commands/'
    PATH_IMAGES = 'images/'
    PATH_IMAGES_DEFAULT = 'images_default/'
    IMAGES_PRODUCT_DEFAULT = 'product.jpg'
    IMAGES_PRODUCTFINAL_DEFAULT = 'productfinal.jpg'
    FILE_IMPORT = 'data/products.xlsx'
    FILE_IMPORT_TMP = 'data/products_tmp.xlsx'
    MSG_PRODUCT_SAMPLE = u'Este producto lo encontrarás en exclusiva en Carrefour España'

    def handle(self, *args, **options):
        # Set debugger
        self.set_debug()
        self.set_name('URBAN LIVING')
        self.debug("Init", color='blue')

        filename_src = self.FILE_IMPORT

        if os.path.isfile(filename_src) is False:
            self.PATH = "{}/{}".format(self.PROJECT, self.PATH)
            filename_src = "{}{}".format(self.PATH, filename_src)

        if os.path.isfile(filename_src) is False:
            self.debug("Fichero no encontrado!!! ({})".format(filename_src), color="red")
        else:
            filename_dst = '{}{}'.format(self.PATH, self.FILE_IMPORT_TMP)

            self.DELETE_ALL()

            # self.regenerate_file(filename_src, filename_dst)
            self.import_excel(filename_src)
            # os.remove(filename_dst)
            self.product_samples(self.MSG_PRODUCT_SAMPLE)
            self.relation_image()
            self.image_default()

    def DELETE_ALL(self):
        self.debug("DELETE ALL ...", color='red')
        OptionValueAttribute.objects.all().delete()
        OptionValueAttribute.objects.all().delete()
        ProductUnique.objects.all().delete()
        ProductFinalAttribute.objects.all().delete()
        ProductFinalImage.objects.all().delete()
        ProductImage.objects.all().delete()
        Attribute.objects.all().delete()
        GroupValueAttribute.objects.all().delete()
        ProductFinal.objects.all().delete()
        Product.objects.all().delete()
        Subcategory.objects.all().delete()
        Category.objects.all().delete()
        Family.objects.all().delete()

    def regenerate_file(self, filename_src, filename_dst):
        self.debug("Regenerate file ...", color='blue')
        wb_new = openpyxl.Workbook()
        name_sheet = wb_new.get_sheet_names()
        sheet_new = wb_new.get_sheet_by_name(name_sheet[0])

        wb = openpyxl.load_workbook(filename_src)
        name_sheet = wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(name_sheet[0])

        last_column = sheet.max_column + 1
        for row in range(2, sheet.max_row + 1):
            self.debug(row, color='cyan')
            description = sheet['F{}'.format(row)].value
            if description:
                name, peso, unit = self.split_description(description)
                for col in range(1, sheet.max_column):
                    cell = "{}{}".format(get_column_letter(col), row)
                    sheet_new[cell] = sheet[cell].value
                cell = "{}{}".format(get_column_letter(last_column), row)
                sheet_new[cell] = name
                cell = "{}{}".format(get_column_letter(last_column + 1), row)
                sheet_new[cell] = peso
                cell = "{}{}".format(get_column_letter(last_column + 2), row)
                sheet_new[cell] = unit

        wb_new.save(filename_dst)

    def split_description(self, description):
        tmp = description.strip().split(' ')
        unit = tmp[-1:][0].upper().replace(')', '')
        peso = tmp[-2:-1][0].upper()
        if unit != 'G' and unit != 'KG':
            if 'KG' in unit:
                peso = unit.split('KG')[0]
                unit = 'KG'
            elif 'G' in unit:
                peso = unit.split('G')[0]
                unit = 'G'
        peso = float(peso.replace(',', '.'))
        name = " ".join(tmp[:-2])
        self.debug(u"{}\t{} -- {} -- {}".format(description, name, peso, unit), color="green")
        return (name, peso, unit)

    def import_excel(self, filename):
        self.debug("Import from excel ...", color='blue')
        # open file
        wb = openpyxl.load_workbook(filename)
        name_sheet = wb.get_sheet_names()

        # select sheet
        sheet = wb.get_sheet_by_name(name_sheet[0])

        self.families = {}
        str_group_value = ['Tipo de envase', 'Unidades', 'Kg']

        for str_group in str_group_value:
            gva = GroupValueAttribute.objects.filter(name=str_group).first()
            if gva is None:
                gva = GroupValueAttribute()
                gva.name = str_group
                gva.save()
            attribute = Attribute.objects.filter(es__description=str_group).first()
            if attribute is None:
                attribute = Attribute()
                attribute.type_value = TYPE_VALUE_LIST
                attribute.list_value = gva
                attribute.order = str_group_value.index(str_group)
                attribute.save()
                for lang_code in settings.LANGUAGES_DATABASES:
                    model = eval("AttributeText{}".format(lang_code))
                    obj = model()
                    obj.attribute = attribute
                    obj.description = str_group
                    obj.save()

        for row in range(2, sheet.max_row + 1):
            self.debug(row, color='cyan')
            # Disponibilidad    Unds o KG Caja  Tipo Envase Vida Útil   Código  Descripción/Ficha Producto  Familia Categoría   Subcategoría    Alérgenos   Ingredientes    Valor energético    Grasas  Grasa saturadas Hidratos de carbono Azúcares    Proteinas   Sal Modo conservación   PVP Unds o KG   PVP Caja    iva no incluido iva no incluido
            # A Disponibilidad
            # B Unds o KG Caja
            # C Tipo Envase ---> atributos
            # D Vida Útil
            # E Código
            # F Descripción comercia
            # G Descripción/Ficha Producto
            # H Peso
            # I Familia
            # J Categoría
            # K Subcategoría
            # L Alérgenos
            # M Ingredientes
            # N Valor energético
            # O Grasas
            # P Grasa saturadas
            # Q Hidratos de carbono
            # R Azúcares
            # S Proteinas
            # T Sal
            # U Modo conservación
            # V PVP Unds o KG
            # W PVP Caja
            # X iva no incluido
            # Y iva no incluido
            if sheet['A{}'.format(row)].value:
                peso = self.get_peso(sheet['H{}'.format(row)].value)
                code = sheet['E{}'.format(row)].value

                family_i = sheet['I{}'.format(row)].value
                category_i = sheet['J{}'.format(row)].value
                subcategory_i = sheet['K{}'.format(row)].value
                tipo_envase = sheet['C{}'.format(row)].value
                unidades_kg = sheet['B{}'.format(row)].value
                peso = self.get_peso(sheet['H{}'.format(row)].value)
                code = sheet['E{}'.format(row)].value

                family = self.get_family(family_i)
                category = self.get_category(family_i, category_i)
                subcategory = self.get_subcategory(family_i, category_i, subcategory_i)

                tax = self.get_tax()

                vida_util = sheet['D{}'.format(row)].value
                price_base = self.get_price_base(tax, sheet['V{}'.format(row)].value, sheet['W{}'.format(row)].value)
                name_product = sheet['F{}'.format(row)].value
                name_product_final = sheet['G{}'.format(row)].value
                alergenos = sheet['L{}'.format(row)].value
                disponibilidad = sheet['A{}'.format(row)].value
                ingredientes = sheet['M{}'.format(row)].value
                valor_energetico = sheet['N{}'.format(row)].value
                grasas = sheet['O{}'.format(row)].value
                grasas_saturadas = sheet['P{}'.format(row)].value
                hidratos = sheet['Q{}'.format(row)].value
                azucar = sheet['R{}'.format(row)].value
                proteinas = sheet['S{}'.format(row)].value
                sal = sheet['T{}'.format(row)].value
                modo_conservacion = sheet['U{}'.format(row)].value
                description_long = self.get_description_long(name_product_final, alergenos, disponibilidad, ingredientes, valor_energetico, grasas, grasas_saturadas, hidratos, azucar, proteinas, sal, modo_conservacion, vida_util)
                description_long_product = self.get_description_long(name_product, alergenos, disponibilidad, ingredientes, valor_energetico, grasas, grasas_saturadas, hidratos, azucar, proteinas, sal, modo_conservacion, vida_util)
                # name_product = sheet['AF{}'.format(row)].value
                self.debug(name_product, color='blue')
                self.debug(name_product_final, color='cyan')
                if name_product is None or name_product == '':
                    name_product = name_product_final

                code_product = nameunify(name_product)
                product = Product.objects.filter(code=code_product, family=family, category=category, subcategory=subcategory).first()
                if not product:
                    product = Product()
                    product.code = code_product
                product.family = family
                product.category = category
                product.subcategory = subcategory
                product.tax = tax
                product.price_base = price_base
                product.force_stock = False
                saved = False
                i = 0
                while not saved:
                    try:
                        product.save()
                        saved = True
                    except IntegrityError:
                        product.code = u"{}{}".format(code_product, i)
                        i += 1

                slug_product = nameunify(name_product)
                for lang_code in settings.LANGUAGES_DATABASES:
                    model = eval("ProductTextText{}".format(lang_code))
                    obj = model.objects.filter(product=product).first()
                    if not obj:
                        obj = model()
                        obj.product = product
                    obj.slug = slug_product
                    obj.name = name_product
                    obj.public = True
                    obj.meta_title = name_product[:70]
                    obj.meta_description = name_product[:70]
                    obj.description_short = name_product
                    obj.description_long = description_long_product
                    saved = False
                    i = 0
                    while not saved:
                        try:
                            obj.save()
                            saved = True
                        except IntegrityError:
                            obj.slug = u"{}{}".format(slug_product, i)
                            i += 1

                pf = ProductFinal.objects.filter(product=product, code=code).first()
                if not pf:
                    pf = ProductFinal()
                    pf.product = product
                    pf.code = code
                pf.price_base_local = price_base
                pf.weight = peso
                saved = False
                i = 0
                while not saved:
                    try:
                        pf.save()
                        saved = True
                    except IntegrityError:
                        pf.code = u"{}{}".format(code, i)
                        i += 1

                slug_product_final = nameunify(name_product_final)
                for lang_code in settings.LANGUAGES_DATABASES:
                    model = eval("ProductFinalText{}".format(lang_code))
                    obj = model.objects.filter(product=pf).first()
                    if not obj:
                        obj = model()
                        obj.product = pf
                    obj.slug = slug_product_final
                    obj.name = name_product_final
                    obj.public = True
                    obj.meta_title = name_product_final[:70]
                    obj.meta_description = name_product_final[:70]
                    obj.description_short = name_product_final
                    obj.description_long = description_long
                    saved = False
                    i = 0
                    while not saved:
                        try:
                            obj.save()
                            saved = True
                        except IntegrityError:
                            obj.slug = u"{}{}".format(slug_product_final, i)
                            i += 1
                gva = GroupValueAttribute.objects.filter(name=str_group_value[0]).first()
                if gva.options_value_attribute.filter(es__description=tipo_envase).exists() is False:
                    opt = OptionValueAttribute()
                    opt.group = gva
                    opt.save()
                    for lang_code in settings.LANGUAGES_DATABASES:
                        model = eval("OptionValueAttributeText{}".format(lang_code))
                        obj = model()
                        obj.option_value = opt
                        obj.description = tipo_envase
                        obj.save()
                else:
                    opt = OptionValueAttribute.objects.filter(es__description=tipo_envase).first()
                attribute = gva.attributes.first()
                pfa = ProductFinalAttribute.objects.filter(product=pf, attribute=attribute).first()
                if not pfa:
                    pfa = ProductFinalAttribute()
                    pfa.product = pf
                    pfa.attribute = attribute
                pfa.value = opt.pk
                pfa.save()

                if 'caja' == tipo_envase.lower():
                    gva_description = str_group_value[2]  # 'Kg'
                    formato = "{} Kg"
                else:
                    gva_description = str_group_value[1]  # 'Unidades'
                    formato = "{} Uds"

                gva = GroupValueAttribute.objects.filter(name=gva_description).first()
                description_op = formato.format(unidades_kg)
                if gva.options_value_attribute.filter(es__description=description_op).exists() is False:
                    op = OptionValueAttribute()
                    op.group = gva
                    op.save()
                    for lang_code in settings.LANGUAGES_DATABASES:
                        model = eval("OptionValueAttributeText{}".format(lang_code))
                        obj = model()
                        obj.option_value = op
                        obj.description = description_op
                        obj.save()
                else:
                    op = OptionValueAttribute.objects.filter(es__description=description_op).first()
                attribute = gva.attributes.first()
                pfa = ProductFinalAttribute.objects.filter(product=pf, attribute=attribute).first()
                if not pfa:
                    pfa = ProductFinalAttribute()
                    pfa.product = pf
                    pfa.attribute = attribute
                pfa.value = op.pk
                pfa.save()

    def get_family(self, family_i):
        # self.debug(family_i, color="white")
        if family_i not in self.families:
            code = nameunify(family_i, True).replace('-', '')
            family = Family.objects.filter(code__istartswith=code).first()
            if family is None:
                family = Family()
                family.code = code
                family.save()

            for lang_code in settings.LANGUAGES_DATABASES:
                model = eval("FamilyText{}".format(lang_code))
                obj = model.objects.filter(family=family).first()
                if obj is None:
                    obj = model()
                obj.family = family
                obj.slug = family_i
                obj.name = family_i
                obj.description = family_i
                obj.save()
            self.families[family_i] = {
                'object': family,
                'categories': {},
            }
        else:
            family = self.families[family_i]['object']
        # self.debug(family.pk, color="red")
        return family

    def get_category(self, family_i, category_i):
        # self.debug(category_i, color="white")
        # self.debug(family_i, color="white")
        if category_i not in self.families[family_i]['categories']:
            code = nameunify(category_i, True).replace('-', '')
            category = Category.objects.filter(code__istartswith=code, family=self.families[family_i]['object']).first()
            if category is None:
                category = Category()
                category.family = self.families[family_i]['object']
                category.code = code
                try:
                    category.save()
                except IntegrityError:
                    category.code = u"{}-{}".format(code, self.families[family_i]['object'].code)
                    try:
                        category.save()
                    except IntegrityError:
                        saved = False
                        i = 0
                        category.code = u"{}{}".format(code, i)
                        while not saved:
                            try:
                                category.save()
                                saved = True
                            except IntegrityError:
                                i += 1
                                category.code = u"{}{}".format(code, i)

                for lang_code in settings.LANGUAGES_DATABASES:
                    model = eval("CategoryText{}".format(lang_code))
                    obj = model()
                    obj.category = category
                    obj.slug = category_i
                    obj.name = category_i
                    obj.description = category_i
                    try:
                        obj.save()
                    except IntegrityError:
                        obj.slug = u"{}-{}".format(category_i, self.families[family_i]['object'].code)
                        try:
                            obj.save()
                        except IntegrityError:
                            saved = False
                            i = 0
                            obj.slug = u'{}{}'.format(category_i, i)
                            while not saved:
                                try:
                                    obj.save()
                                    saved = True
                                except IntegrityError:
                                    i += 1
                                    obj.slug = u'{}{}'.format(category_i, i)

            self.families[family_i]['categories'][category_i] = {
                'object': category,
                'subcategories': {},
            }
        else:
            category = self.families[family_i]['categories'][category_i]['object']
        # self.debug(category.pk, color="red")
        return category

    def get_subcategory(self, family_i, category_i, subcategory_i):
        if subcategory_i is None:
            self.debug("subcategory_i is None", color="simplered")
            subcategory_i = category_i

        # self.debug(subcategory_i, color="white")
        if subcategory_i not in self.families[family_i]['categories'][category_i]['subcategories']:
            code = nameunify(subcategory_i, True).replace('-', '')
            subcategory = Subcategory.objects.filter(code__istartswith=code, category=self.families[family_i]['categories'][category_i]['object']).first()
            if subcategory is None:
                subcategory = Subcategory()
                subcategory.category = self.families[family_i]['categories'][category_i]['object']
                subcategory.code = code
                try:
                    subcategory.save()
                except IntegrityError:
                    subcategory.code = u'{}-{}'.format(code, self.families[family_i]['categories'][category_i]['object'].code)
                    try:
                        subcategory.save()
                    except IntegrityError:
                        subcategory.code = u'{}-{}'.format(code, self.families[family_i]['object'].code)
                        try:
                            subcategory.save()
                        except IntegrityError:
                            subcategory.code = u'{}-{}-{}'.format(code, self.families[family_i]['categories'][category_i]['object'], self.families[family_i]['object'].code)
                            try:
                                subcategory.save()
                            except IntegrityError:
                                saved = False
                                i = 0
                                subcategory.code = u"{}{}".format(code, i)
                                while not saved:
                                    try:
                                        subcategory.save()
                                        saved = True
                                    except IntegrityError:
                                        i += 1
                                        subcategory.code = u"{}{}".format(code, i)

                for lang_code in settings.LANGUAGES_DATABASES:
                    model = eval("SubcategoryText{}".format(lang_code))
                    obj = model()
                    obj.subcategory = subcategory
                    obj.slug = subcategory.code
                    obj.name = subcategory_i
                    obj.description = subcategory_i
                    saved = False
                    i = 0
                    while not saved:
                        try:
                            obj.save()
                            saved = True
                        except IntegrityError:
                            obj.slug = u'{}{}'.format(subcategory.code, i)
                            i += 1
        else:
            subcategory = self.families[family_i]['categories'][category_i]['subcategories']['object']
        # self.debug(subcategory.pk, color="red")
        return subcategory

    def get_tax(self):
        code = '21'
        # self.debug(code, color="white")
        tax = TypeTax.objects.filter(name=code).first()
        if not tax:
            tax = TypeTax()
            tax.name = code
            tax.tax = int(code)
            tax.recargo_equivalencia = 0
            tax.save()

        # self.debug(tax.pk, color="red")
        return tax

    def get_price_base(self, tax, price_unit, price_box):
        if price_unit:
            price = price_unit
        else:
            price = price_box

        self.debug("PRICE {}".format(price), color="yellow")
        try:
            price = float(price)
        except ValueError:
            self.debug("PRICE INVALID {}".format(price), color="red")
            price = 0

        price_base = price / (1 + (tax.tax / 100))
        # self.debug("{} -- {} -- {}".format(price, price_base, tax.tax), color="green")
        return price_base

    def get_peso(self, peso):
        peso = str(peso).lower().replace(',', '.')
        tmp = peso.split(' ')
        if len(tmp) > 1:
            new_peso = tmp[0]
            if 'k' in tmp[1]:
                unidad = 1000
            else:
                unidad = 1
        else:
            if 'k' in tmp[0]:
                tmp = tmp[0].split('k')
                unidad = 1000
            else:
                tmp = tmp[0].split('g')
                unidad = 1
            new_peso = tmp[0]

        new_peso = float(new_peso) * unidad

        return new_peso

    def get_description_long(self, description, alergenos, disponibilidad, ingredientes, valor_energetico, grasas, grasas_saturadas, hidratos, azucar, proteinas, sal, modo_conservacion, vida_util):
        text = u"""
        <div class="ingredientes_product">
        <h3>Ingredientes</h3>
        <p>{ingredientes}</p>
        </div>

        <div class="nutricional_product">
        <h3>Información nutricional</h3>
        <table>
            <thead>
            <tr>
                <td>Valor energético</td>
                <td>Grasas</td>
                <td>Grasas saturadas</td>
                <td>Hidratos</td>
                <td>Azúcar</td>
                <td>Proteínas</td>
                <td>Sal</td>
            </tr>
            </thead>
            <tbody>
            <tr>
                <td>{valor_energetico}</td>
                <td>{grasas}</td>
                <td>{grasas_saturadas}</td>
                <td>{hidratos}</td>
                <td>{azucar}</td>
                <td>{proteinas}</td>
                <td>{sal}</td>
            </tr>
            </tbody>
        </table>
        </div>

        <div class="alergenos_product">
        <h3>Alérgenos</h3>
        <p>{alergenos}</p>
        </div>

        <div class="vida_util_product">
        <h3>Vida útil</h3>
        <p>{vida_util}</p>
        </div>

        <div class="vida_conservacion">
        <h3>Modo de conservación</h3>
        <p>{modo_conservacion}</p>
        </div>

        <div class="vida_disponibilidad">
        <h3>Disponibilidad</h3>
        <p>{disponibilidad}</p>
        </div>
        """.format(**{
            'description': description,
            'alergenos': alergenos,
            'ingredientes': ingredientes,
            'valor_energetico': valor_energetico,
            'grasas': grasas,
            'grasas_saturadas': grasas_saturadas,
            'hidratos': hidratos,
            'azucar': azucar,
            'proteinas': proteinas,
            'sal': sal,
            'disponibilidad': disponibilidad,
            'modo_conservacion': modo_conservacion,
            'vida_util': vida_util,
        })
        return text

    def product_samples(self, msg):
        self.debug(u"Todos los productos de la familia ORIGN 1948 son de muestras..", color="green")
        family = Family.objects.filter(es__name__icontains='1948').first()
        for pf in ProductFinal.objects.filter(product__family=family):
            pf.sample = True
            pf.save()
            for lang_code in settings.LANGUAGES_DATABASES:
                model = eval("ProductFinalText{}".format(lang_code))
                obj = model.objects.filter(product=pf).first()
                obj.description_sample = msg
                obj.save()

    def image_default(self):
        self.debug(u"Asignando imagen por defecto a productos y productos finales sin imagenes", color="green")
        image_product = '{}{}{}'.format(self.PATH, self.PATH_IMAGES_DEFAULT, self.IMAGES_PRODUCT_DEFAULT)
        FILE = File(open(image_product, 'rb'))
        # Prepare file in memory
        MEM = io.BytesIO()
        MEM.write(FILE.read())
        MEM.name = self.IMAGES_PRODUCT_DEFAULT
        MEM.seek(0)
        FILE.close()

        queryset = Product.objects.filter(products_image__isnull=True)
        for product in queryset:
            pi = ProductImage()
            pi.product = product
            pi.public = True
            pi.principal = True
            pi.image = File(MEM)
            pi.name_file = self.IMAGES_PRODUCT_DEFAULT
            pi.save()
        self.debug(u"Resumen: --> Productos {}".format(queryset.count()), color="blue")

        image_product_final = '{}{}{}'.format(self.PATH, self.PATH_IMAGES_DEFAULT, self.IMAGES_PRODUCTFINAL_DEFAULT)
        FILE = File(open(image_product_final, 'rb'))
        # Prepare file in memory
        MEM = io.BytesIO()
        MEM.write(FILE.read())
        MEM.name = self.IMAGES_PRODUCTFINAL_DEFAULT
        MEM.seek(0)
        FILE.close()

        queryset = ProductFinal.objects.filter(productfinals_image__isnull=True)
        for product in queryset:
            pi = ProductFinalImage()
            pi.product_final = product
            pi.public = True
            pi.image = File(MEM)
            pi.name_file = self.IMAGES_PRODUCTFINAL_DEFAULT
            pi.save()
        self.debug(u"Resumen: --> Productos {}".format(queryset.count()), color="blue")

    def relation_image(self):
        self.debug(u"Relacionando imagenes de productos finales", color="green")
        path = '{}{}'.format(self.PATH, self.PATH_IMAGES)
        total = 0
        new = 0
        update = 0
        notexists = 0
        for arch in listdir(path):
            filename = join(path, arch)
            if isfile(join(path, arch)):
                total += 1
                codes = arch.split('.')[0].split('_')[0].split('-')
                code = codes[0].split("+")[0]
                code = code.split(" ")[0]
                if len(codes) != 1:
                    image_product_final = False
                else:
                    image_product_final = True

                self.debug("{}\t{}".format(code, arch), color="green")
                pf = ProductFinal.objects.filter(code=code).first()
                if pf is None:
                    notexists += 1
                else:
                    FILE = File(open(filename, 'rb'))
                    # Prepare file in memory
                    MEM = io.BytesIO()
                    MEM.write(FILE.read())
                    MEM.name = arch
                    MEM.seek(0)
                    FILE.close()
                    if image_product_final:
                        pfi = ProductFinalImage.objects.filter(product_final=pf, name_file=filename).first()
                        text_model = "ProductFinalImageText"
                    else:
                        pfi = ProductImage.objects.filter(product=pf.product, name_file=filename).first()
                        text_model = "ProductImageText"

                    if pfi is None:
                        new += 1
                        if image_product_final:
                            img_tmp = ProductFinalImage.objects.filter(product_final=pf, name_file=self.IMAGES_PRODUCTFINAL_DEFAULT).first()
                            if img_tmp:
                                img_tmp.delete()
                            pfi = ProductFinalImage()
                            pfi.product_final = pf
                        else:
                            img_tmp = ProductImage.objects.filter(product=pf.product, name_file=self.IMAGES_PRODUCTFINAL_DEFAULT).first()
                            if img_tmp:
                                img_tmp.delete()
                            pfi = ProductImage()
                            pfi.product = pf.product
                    else:
                        update += 1

                    pfi.image = File(MEM)
                    pfi.name_file = filename
                    pfi.principal = True
                    pfi.save()

                    file_description = pf.es.name

                    for lang_code in settings.LANGUAGES_DATABASES:
                        model = eval("{}{}".format(text_model, lang_code))
                        if image_product_final:
                            obj = model.objects.filter(product_final_image=pfi).first()
                        else:
                            obj = model.objects.filter(product_image=pfi).first()
                        if obj is None:
                            obj = model()
                        obj.name = file_description
                        obj.description = file_description
                        if image_product_final:
                            obj.product_final_image = pfi
                        else:
                            obj.product_image = pfi
                        obj.save()

        self.debug(u"Resumen: --> Total: {}".format(total), color="blue")
        self.debug(u"Resumen: --> Nuevos: {}".format(new), color="blue")
        self.debug(u"Resumen: --> Actualizados: {}".format(update), color="blue")
        self.debug(u"Resumen: --> No existen: {}".format(notexists), color="red")
