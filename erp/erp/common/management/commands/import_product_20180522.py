# -*- coding: utf-8 -*-
import decimal

import io
import os
from os import listdir
from os.path import isfile, join
import openpyxl
from openpyxl.utils import get_column_letter

from django.db import IntegrityError
from django.core.files import File
from django.core.management.base import BaseCommand
from django.conf import settings
from django.utils.translation import ugettext_lazy as _

from codenerix.helpers import nameunify
from codenerix.lib.debugger import Debugger
from codenerix_products.models import (
    TypeTax,
    Family,
    Category,
    Subcategory,
    MODELS_SLUG,
    MODELS_BRANDS,
    MODELS_PRODUCTS,
    MODELS_PRODUCTS_FINAL,
    Product,
    ProductFinal,
    Attribute,
    MODELS,
    TYPE_VALUE_LIST,
    GroupValueAttribute,
    OptionValueAttribute,
    ProductFinalImage,
    ProductFinalAttribute,
    ProductImage,
    ProductUnique,
)
from codenerix_products.models import GroupValueFeature, Feature

from codenerix_storages.models import (
    StorageBox,
    StorageBoxKind,
    StorageBoxStructure,
    StorageZone,
    Storage,
)

for info in (
    MODELS_SLUG + MODELS_BRANDS + MODELS_PRODUCTS + MODELS + MODELS_PRODUCTS_FINAL
):
    field = info[0]
    model = info[1]
    for lang_code in settings.LANGUAGES_DATABASES:
        query = "from codenerix_products.models import {}Text{}\n".format(
            model, lang_code
        )
        exec(query)


class Command(BaseCommand, Debugger):
    FILE_IMPORT_FAMILY = "data/products_20180522_1.xlsx"
    FILE_IMPORT_PRODUCT = "data/products_20180522_2.xlsx"
    FILE_IMPORT_STOCK = "data/products_20180522_3.xlsx"
    PATH = "common/management/commands/"
    PROJECT = "erp"

    IMAGES_PRODUCT_DEFAULT = "product.jpg"
    IMAGES_PRODUCTFINAL_DEFAULT = "productfinal.jpg"
    FILE_IMPORT_TMP = "data/products_tmp.xlsx"

    def handle(self, *args, **options):
        # Set debugger
        self.set_debug()
        self.set_name("Codenerix")
        self.debug("Init", color="blue")

        self.DELETE_ALL()

        filename_src = self.FILE_IMPORT_FAMILY
        if os.path.isfile(filename_src) is False:
            path = "{}/{}".format(self.PROJECT, self.PATH)
            filename_src = "{}{}".format(path, filename_src)
        # """
        if os.path.isfile(filename_src) is False:
            self.debug(
                "Fichero no encontrado!!! ({})".format(filename_src), color="red"
            )
        else:
            self.families_categories_subcategories(filename_src)
        # """
        filename_src = self.FILE_IMPORT_PRODUCT
        if os.path.isfile(filename_src) is False:
            path = "{}/{}".format(self.PROJECT, self.PATH)
            filename_src = "{}{}".format(path, filename_src)

        # """
        if os.path.isfile(filename_src) is False:
            self.debug(
                "Fichero no encontrado!!! ({})".format(filename_src), color="red"
            )
        else:
            self.products(filename_src)
        # """
        filename_src = self.FILE_IMPORT_STOCK
        if os.path.isfile(filename_src) is False:
            path = "{}/{}".format(self.PROJECT, self.PATH)
            filename_src = "{}{}".format(path, filename_src)

        if os.path.isfile(filename_src) is False:
            self.debug(
                "Fichero no encontrado!!! ({})".format(filename_src), color="red"
            )
        else:
            self.stocks(filename_src)

    def DELETE_ALL(self):
        self.debug("DELETE ALL ...", color="red")
        self.debug("Delete OptionValueAttribute", color="simplered")
        OptionValueAttribute.objects.all().delete()
        self.debug("Delete OptionValueAttribute", color="simplered")
        OptionValueAttribute.objects.all().delete()
        self.debug("Delete ProductUnique", color="simplered")
        ProductUnique.objects.all().delete()
        self.debug("Delete ProductFinalAttribute", color="simplered")
        ProductFinalAttribute.objects.all().delete()
        self.debug("Delete ProductFinalImage", color="simplered")
        ProductFinalImage.objects.all().delete()
        self.debug("Delete ProductImage", color="simplered")
        ProductImage.objects.all().delete()
        self.debug("Delete Attribute", color="simplered")
        Attribute.objects.all().delete()
        self.debug("Delete GroupValueAttribute", color="simplered")
        GroupValueAttribute.objects.all().delete()
        self.debug("Delete ProductFinal", color="simplered")
        ProductFinal.objects.all().delete()
        self.debug("Delete Product", color="simplered")
        Product.objects.all().delete()
        self.debug("Delete Subcategory", color="simplered")
        Subcategory.objects.all().delete()
        self.debug("Delete Category", color="simplered")
        Category.objects.all().delete()
        self.debug("Delete Family", color="simplered")
        Family.objects.all().delete()

    def stocks(self, filename):
        self.debug("Import from excel ...", color="blue")
        box = self.get_box()
        # open file
        wb = openpyxl.load_workbook(filename)
        self.debug("load_workbook ...", color="blue")
        name_sheet = wb.get_sheet_names()
        self.debug("get_sheet_names ...", color="blue")

        # select sheet
        sheet = wb.get_sheet_by_name(name_sheet[0])
        self.debug("get_sheet_by_name Family ...", color="blue")

        self.families = {}
        self.categories = {}

        for row in range(3, sheet.max_row + 1):
            # A CÓDIGOS EAN 13
            # B DESCRIPCIÓN
            # C CÓDIGO INTERNO
            # D COLOR
            # E TALLA
            # F PVP
            # G Precio Coste
            # H Artículo
            # I Continuidad
            # J Temporada
            # K Stock
            if sheet["A{}".format(row)].value:
                ean13 = sheet["A{}".format(row)].value
                code_product_final = sheet["C{}".format(row)].value
                color = sheet["D{}".format(row)].value
                talla = sheet["E{}".format(row)].value
                stock = sheet["K{}".format(row)].value

                self.debug("{} - {}".format(ean13, code_product_final), color="red")

                if isinstance(ean13, float):
                    ean13 = int(ean13)
                ean13 = str(ean13).strip()

                if isinstance(code_product_final, float):
                    code_product_final = int(code_product_final)
                code_product_final = str(code_product_final).strip()

                self.debug("{} - {}".format(ean13, code_product_final), color="green")
                pf = ProductFinal.objects.filter(ean13=ean13).first()
                if pf:
                    pu = ProductUnique()
                    pu.product_final = pf
                    pu.box = box
                    pu.stock_original = stock
                    pu.stock_real = stock
                    pu.save()

    def products(self, filename):

        group_color, attribute_color = self.get_attributes("color")
        group_talla, attribute_talla = self.get_attributes("talla")

        tax = TypeTax.objects.filter(default=True).first()

        self.debug("Import from excel ...", color="blue")
        # open file
        wb = openpyxl.load_workbook(filename)
        self.debug("load_workbook ...", color="blue")
        name_sheet = wb.get_sheet_names()
        self.debug("get_sheet_names ...", color="blue")

        # select sheet
        sheet = wb.get_sheet_by_name(name_sheet[0])
        self.debug("get_sheet_by_name Product ...", color="blue")
        for row in range(3, sheet.max_row + 1):
            # A Id – codigo producto
            # B Referencia – codigo final
            # C descripcion
            # D descripcion_larga
            # E FAMILIA
            # F MODELO
            # G COLECCION
            # H CATEGORIA
            # I SUBCATEGORIA
            # J nombre_talla
            # K nombre_Color
            # L Codigobarras – ean13
            # M stock
            # N stockminimo
            # O tarifa_pvp
            # P tarifa_coste
            # Q tarifa_multimarca
            # R tarifa_pmp_grupo
            # S tarifa_retail
            # T tarifa_pmp_empresa
            # U ERROR

            if sheet["A{}".format(row)].value:
                code_product = sheet["A{}".format(row)].value
                code_producf_final = sheet["B{}".format(row)].value
                name_product = sheet["C{}".format(row)].value
                description_long_product = sheet["D{}".format(row)].value
                family = sheet["E{}".format(row)].value
                model = sheet["F{}".format(row)].value
                COLECCION = sheet["G{}".format(row)].value
                category = sheet["H{}".format(row)].value
                subcategory = sheet["I{}".format(row)].value
                nombre_talla = sheet["J{}".format(row)].value
                nombre_color = sheet["K{}".format(row)].value
                ean13 = sheet["L{}".format(row)].value
                tarifa_pvp = decimal.Decimal(str(sheet["O{}".format(row)].value))

                if isinstance(code_producf_final, float):
                    code_producf_final = int(code_producf_final)
                code_producf_final = str(code_producf_final).strip()

                if isinstance(code_product, float):
                    code_product = int(code_product)
                code_product = str(code_product).strip()

                if isinstance(ean13, float):
                    ean13 = int(ean13)
                ean13 = str(ean13).strip()

                family_obj = Family.objects.filter(es__name=family).first()
                if family_obj is None:
                    self.debug("Familia no encontrada. {}".format(family), color="red")
                    raise

                if category == "C/CAJA":
                    self.debug(
                        "Categoria truncada. {}".format(category), color="yellow"
                    )
                    category = "C/ CAJA"
                if category == "BAÑADORES ESTAMPADO":
                    self.debug(
                        "Categoria truncada. {}".format(category), color="yellow"
                    )
                    category = "BAÑADOR ESTAMPADO"
                if category == "BAÑADORES LISO":
                    self.debug(
                        "Categoria truncada. {}".format(category), color="yellow"
                    )
                    category = "BAÑADOR LISO"

                category_obj = Category.objects.filter(
                    es__name=category, family=family_obj
                ).first()
                if category == "BERMUDA CHINO SPORT":
                    self.debug(
                        "Categoria no encontrada. {}".format(category), color="red"
                    )
                elif category_obj is None:
                    self.debug(
                        "Categoria no encontrada. {}".format(category), color="red"
                    )
                    raise

                if category_obj is None:
                    continue

                if subcategory is None:
                    subcategory = category
                    self.debug(
                        "Subcategoria truncada. {}".format(category), color="yellow"
                    )
                subcategory_obj = None
                if subcategory:
                    subcategory_obj = Subcategory.objects.filter(
                        es__name=subcategory, category=category_obj
                    ).first()
                    if (
                        subcategory_obj is None
                        and subcategory != "TRAJES CON CHALECO LISO"
                    ):
                        self.debug(
                            "Subcategoria no encontrada. {}".format(subcategory),
                            color="yellow",
                        )

                if subcategory_obj is None:
                    self.debug(
                        "Creando subcategoria. {}".format(category), color="cyan"
                    )
                    subcategory_obj = self.get_subcategory(
                        category_obj.code, subcategory, subcategory
                    )

                # ##############
                if subcategory_obj is None:
                    self.debug(
                        "Subcategoria no encontrada!!!! {}".format(subcategory),
                        color="yellow",
                    )
                    raise
                else:

                    product = Product.objects.filter(
                        code=code_product,
                        family=family_obj,
                        category=category_obj,
                        subcategory=subcategory_obj,
                    ).first()
                    if not product:
                        product = Product()
                        product.code = code_product
                    product.family = family_obj
                    product.category = category_obj
                    product.subcategory = subcategory_obj
                    product.tax = tax
                    product.price_base = tarifa_pvp
                    product.force_stock = False
                    saved = False
                    i = 0
                    while not saved:
                        try:
                            product.save()
                            saved = True
                        except IntegrityError:
                            product.code = "{}{}".format(code_product, i)
                            i += 1

                    slug_product = nameunify(code_product)
                    for lang_code in settings.LANGUAGES_DATABASES:
                        model = eval("ProductTextText{}".format(lang_code))
                        obj = model.objects.filter(product=product).first()
                        if not obj:
                            obj = model()
                            obj.product = product
                        obj.slug = slug_product
                        obj.name = name_product
                        obj.public = True
                        obj.meta_title = name_product[:70]
                        obj.meta_description = name_product[:70]
                        obj.description_short = name_product
                        obj.description_long = description_long_product
                        saved = False
                        i = 0
                        while not saved:
                            try:
                                obj.save()
                                saved = True
                            except IntegrityError:
                                obj.slug = "{}{}".format(slug_product, i)
                                i += 1

                    pf = ProductFinal.objects.filter(
                        product=product, code=code_producf_final
                    ).first()
                    if not pf:
                        pf = ProductFinal()
                        pf.product = product
                        pf.code_producf_final = code_producf_final
                    pf.price_base_local = tarifa_pvp
                    pf.ean13 = ean13
                    saved = False
                    i = 0
                    while not saved:
                        try:
                            pf.save()
                            saved = True
                        except IntegrityError:
                            pf.code = "{}{}".format(code_producf_final, i)
                            i += 1

                    slug_product_final = nameunify(name_product)
                    for lang_code in settings.LANGUAGES_DATABASES:
                        model = eval("ProductFinalText{}".format(lang_code))
                        obj = model.objects.filter(product=pf).first()
                        if not obj:
                            obj = model()
                            obj.product = pf
                        obj.slug = slug_product_final
                        obj.name = name_product
                        obj.public = True
                        obj.meta_title = name_product[:70]
                        obj.meta_description = name_product[:70]
                        obj.description_short = name_product
                        obj.description_long = description_long_product
                        saved = False
                        i = 0
                        while not saved:
                            try:
                                obj.save()
                                saved = True
                            except IntegrityError:
                                obj.slug = "{}{}".format(slug_product_final, i)
                                i += 1

                    self.debug("Product: {} \t\t\tProductFinal: {}".format(product, pf))

                    # ###########

                    if (
                        group_color.options_value_attribute.filter(
                            es__description=nombre_color
                        ).exists()
                        is False
                    ):
                        op = OptionValueAttribute()
                        op.group = group_color
                        op.save()
                        for lang_code in settings.LANGUAGES_DATABASES:
                            model = eval("OptionValueAttributeText{}".format(lang_code))
                            obj = model()
                            obj.option_value = op
                            obj.description = nombre_color
                            obj.save()
                    else:
                        op = OptionValueAttribute.objects.filter(
                            es__description=nombre_color
                        ).first()

                    pfa = ProductFinalAttribute.objects.filter(
                        product=pf, attribute=attribute_color
                    ).first()
                    if not pfa:
                        pfa = ProductFinalAttribute()
                        pfa.product = pf
                        pfa.attribute = attribute_color
                    pfa.value = op.pk
                    pfa.save()

                    if (
                        group_talla.options_value_attribute.filter(
                            es__description=nombre_talla
                        ).exists()
                        is False
                    ):
                        op = OptionValueAttribute()
                        op.group = group_talla
                        op.save()
                        for lang_code in settings.LANGUAGES_DATABASES:
                            model = eval("OptionValueAttributeText{}".format(lang_code))
                            obj = model()
                            obj.option_value = op
                            obj.description = nombre_talla
                            obj.save()
                    else:
                        op = OptionValueAttribute.objects.filter(
                            es__description=nombre_talla
                        ).first()

                    pfa = ProductFinalAttribute.objects.filter(
                        product=pf, attribute=attribute_talla
                    ).first()
                    if not pfa:
                        pfa = ProductFinalAttribute()
                        pfa.product = pf
                        pfa.attribute = attribute_talla
                    pfa.value = op.pk
                    pfa.save()

    def families_categories_subcategories(self, filename):
        self.debug("Import from excel ...", color="blue")
        # open file
        wb = openpyxl.load_workbook(filename)
        self.debug("load_workbook ...", color="blue")
        name_sheet = wb.get_sheet_names()
        self.debug("get_sheet_names ...", color="blue")

        # select sheet
        sheet = wb.get_sheet_by_name(name_sheet[0])
        self.debug("get_sheet_by_name Family ...", color="blue")

        self.families = {}
        self.categories = {}
        for row in range(3, sheet.max_row + 1):
            # A Código
            # B Orden
            # C Público
            # D Mostrar en frontend
            #  ESPAÑOL
            # E Nombre
            # F URL Amigable
            # G Descripción
            # H Meta title
            # I Meta description
            # J Meta keywords
            #  INGLES
            # K Nombre
            # L URL Amigable
            # M Descripción Meta title
            # N Meta description
            # O Meta keywords
            if sheet["A{}".format(row)].value:
                self.debug(row, color="cyan")
                code = sheet["A{}".format(row)].value
                name = sheet["E{}".format(row)].value
                family = self.get_family(str(int(code)).strip(), name)

        sheet = wb.get_sheet_by_name(name_sheet[1])
        self.debug("get_sheet_by_name Category ...", color="blue")
        for row in range(3, sheet.max_row + 1):
            # A Codigo
            # B Orden
            # C Público
            # D Mostrar en frontend
            # E Familia
            # F Mostrar en el frontend productos con stock
            #  ESPAÑOL
            # G Nombre
            # H URL Amigable
            # I Descripción
            # J Meta title
            # K Meta description
            # L Meta keywords
            #  INGLES
            # M Nombre
            # N URL Amigable
            # O Descripción
            # P Meta title
            # Q Meta description
            # R Meta keywords
            if sheet["A{}".format(row)].value:
                self.debug(row, color="cyan")
                code = sheet["A{}".format(row)].value
                code_family = sheet["E{}".format(row)].value
                name = sheet["G{}".format(row)].value
                category = self.get_category(
                    str(int(code_family)).strip(), str(int(code)).strip(), name
                )

        sheet = wb.get_sheet_by_name(name_sheet[2])
        self.debug("get_sheet_by_name Subcategory ...", color="blue")
        for row in range(3, sheet.max_row + 1):
            self.debug(row, color="cyan")
            # A Codigo
            # B Orden
            # C Público
            # D Mostrar en frontend
            # E Categoría
            # F Mostrar la marca (frontend
            # G Destacado
            #  ESPAÑOL
            # H Nombre
            # I URL Amigable
            # J Descripción
            # K Meta title
            # L Meta description
            # M Meta keywords
            #  INGLES
            # N Nombre
            # O URL Amigable
            # P Descripción
            # Q Meta title
            # R Meta description
            # S Meta keywords
            if sheet["A{}".format(row)].value:
                self.debug(row, color="cyan")
                code = sheet["A{}".format(row)].value
                code_category = sheet["E{}".format(row)].value
                name = sheet["H{}".format(row)].value
                if code and name and code_category:
                    subcategory = self.get_subcategory(
                        str(int(code_category)).strip(), code, name
                    )

        sheet = wb.get_sheet_by_name(name_sheet[4])
        self.debug("get_sheet_by_name TypeTax ...", color="blue")
        for row in range(3, sheet.max_row + 1):
            # A Nombre
            # B Porcentaje
            # C Impuesto por defecto
            # D recargo de equivalencia
            if sheet["A{}".format(row)].value:
                name = sheet["A{}".format(row)].value
                self.debug(row, color="cyan")
                tax = sheet["B{}".format(row)].value
                default = sheet["C{}".format(row)].value
                recargo = sheet["D{}".format(row)].value
                tax = self.get_tax(
                    name.strip(), float(tax), bool(default), float(recargo)
                )

    def get_family(self, code, name):
        if code and name:
            if code not in self.families:
                family = Family.objects.filter(code=code).first()
                if family is None:
                    family = Family()
                    family.code = code
                    family.save()

                    for lang_code in settings.LANGUAGES_DATABASES:
                        model = eval("FamilyText{}".format(lang_code))
                        obj = model.objects.filter(family=family).first()
                        if obj is None:
                            obj = model()
                        obj.family = family
                        obj.slug = name
                        obj.name = name
                        obj.description = name
                        obj.meta_title = name
                        obj.meta_description = name
                        obj.meta_keywords = name
                        obj.save()
                self.families[code] = {
                    "object": family,
                    "categories": {},
                }
            else:
                family = self.families[code]["object"]
        else:
            family = None
        return Family

    def get_category(self, code_family, code, name):
        # self.debug(category_i, color="white")
        # self.debug(family_i, color="white")
        if code_family and code and name:
            if code not in self.families[code_family]["categories"]:
                category = Category.objects.filter(
                    code=code, family=self.families[code_family]["object"]
                ).first()
                if category is None:
                    category = Category()
                    category.family = self.families[code_family]["object"]
                    category.code = code
                    try:
                        category.save()
                    except IntegrityError:
                        category.code = "{}-{}".format(
                            code, self.families[code_family]["object"].code
                        )
                        try:
                            category.save()
                        except IntegrityError:
                            saved = False
                            i = 0
                            category.code = "{}{}".format(code, i)
                            while not saved:
                                try:
                                    category.save()
                                    saved = True
                                except IntegrityError:
                                    i += 1
                                    category.code = "{}{}".format(code, i)

                    for lang_code in settings.LANGUAGES_DATABASES:
                        model = eval("CategoryText{}".format(lang_code))
                        obj = model()
                        obj.category = category
                        obj.slug = name
                        obj.name = name
                        obj.description = name
                        try:
                            obj.save()
                        except IntegrityError:
                            obj.slug = "{}-{}".format(
                                name, self.families[code_family]["object"].code
                            )
                            try:
                                obj.save()
                            except IntegrityError:
                                saved = False
                                i = 0
                                obj.slug = "{}{}".format(name, i)
                                while not saved:
                                    try:
                                        obj.save()
                                        saved = True
                                    except IntegrityError:
                                        i += 1
                                        obj.slug = "{}{}".format(name, i)

                self.families[code_family]["categories"][code] = {
                    "object": category,
                    "subcategories": {},
                }
                self.categories[code] = self.families[code_family]["categories"][code]
            else:
                category = self.families[code_family]["categories"][code]["object"]
                self.categories[code] = category
        else:
            category = None
        return category

    def get_subcategory(self, code_category, code, name):
        code = nameunify(name, True).replace("-", "")
        if code_category and code and name:

            # self.debug(subcategory_i, color="white")
            if code_category not in self.categories[code_category]["subcategories"]:
                category = self.categories[code_category]["object"]
                subcategory = Subcategory.objects.filter(
                    code=code, category=category
                ).first()
                if subcategory is None:
                    subcategory = Subcategory()
                    subcategory.category = category
                    subcategory.code = code
                    try:
                        subcategory.save()
                    except IntegrityError:
                        subcategory.code = "{}-{}".format(code, category.code)
                        try:
                            subcategory.save()
                        except IntegrityError:
                            subcategory.code = "{}-{}".format(
                                code, category.family.code
                            )
                            try:
                                subcategory.save()
                            except IntegrityError:
                                subcategory.code = "{}-{}-{}".format(
                                    code, category.code, category.family.code
                                )
                                try:
                                    subcategory.save()
                                except IntegrityError:
                                    saved = False
                                    i = 0
                                    subcategory.code = "{}{}".format(code, i)
                                    while not saved:
                                        try:
                                            subcategory.save()
                                            saved = True
                                        except IntegrityError:
                                            i += 1
                                            subcategory.code = "{}{}".format(code, i)

                    for lang_code in settings.LANGUAGES_DATABASES:
                        model = eval("SubcategoryText{}".format(lang_code))
                        obj = model()
                        obj.subcategory = subcategory
                        obj.slug = subcategory.code
                        obj.name = name
                        obj.description = name
                        saved = False
                        i = 0
                        while not saved:
                            try:
                                obj.save()
                                saved = True
                            except IntegrityError:
                                obj.slug = "{}{}".format(subcategory.code, i)
                                i += 1
            else:
                subcategory = self.families[category.family.code]["categories"][
                    code_category
                ]["subcategories"]["object"]
        else:
            subcategory = None

        return subcategory

    def get_tax(self, name, tax_value, default, recargo):
        tax = TypeTax.objects.filter(name=name).first()
        if not tax:
            tax = TypeTax()
            tax.name = name
            tax.tax = tax_value
            tax.default = default
            tax.recargo_equivalencia = recargo
            tax.save()

        # self.debug(tax.pk, color="red")
        return tax

    def get_attributes(self, name):

        gva = GroupValueAttribute.objects.filter(name=name).first()
        if gva is None:
            gva = GroupValueAttribute()
            gva.name = name
            gva.save()
        attribute = Attribute.objects.filter(es__description=name).first()
        if attribute is None:
            attribute = Attribute()
            attribute.type_value = TYPE_VALUE_LIST
            attribute.list_value = gva
            attribute.order = 1
            attribute.save()
            for lang_code in settings.LANGUAGES_DATABASES:
                model = eval("AttributeText{}".format(lang_code))
                obj = model()
                obj.attribute = attribute
                obj.description = name
                obj.save()

        return [gva, attribute]

    def get_box(self):
        box = StorageBox.objects.first()
        if box is None:
            box = StorageBox()
            box.box_kind = self.get_box_kind()
            box.box_structure = self.get_box_structure()
            box.name = "Caja prueba"
            box.save()
        return box

    def get_box_kind(self):
        box_kind = StorageBoxKind.objects.first()
        if box_kind is None:
            box_kind = StorageBoxKind()
            box_kind.name = "Box Kind prueba"
            box_kind.save()
        return box_kind

    def get_box_structure(self):
        box_structure = StorageBoxStructure.objects.first()
        if box_structure is None:
            box_structure = StorageBoxStructure()
            box_structure.zone = self.get_box_zone()
            box_structure.name = "Box Kind prueba"
            box_structure.save()
        return box_structure

    def get_box_zone(self):
        box_zone = StorageZone.objects.first()
        if box_zone is None:
            box_zone = StorageZone()
            box_zone.storage = self.get_storage()
            box_zone.name = "Box Kind prueba"
            box_zone.save()
        return box_zone

    def get_storage(self):
        storage = Storage.objects.first()
        if storage is None:
            storage = Storage()
            storage.name = "Box Kind prueba"
            storage.save()
        return storage
